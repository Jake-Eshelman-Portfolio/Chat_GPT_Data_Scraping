import subprocess
import time
import os
import json
from openai import OpenAI
import openpyxl
import tkinter as tk
from tkinter import messagebox
import sys

config_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'config'))
sys.path.append(config_path)

import config

client = OpenAI(
    api_key=config.OPENAI_API_KEY
)

def extract_json_from_response(response):
    start = response.find('[')
    end = response.rfind(']') + 1
    if start != -1 and end != -1:
        return response[start:end]
    return None

def get_chatgpt_response(prompt, model):
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"Error: {str(e)}"

# Sanitize the title to make it a valid filename
def generate_file_name(title):
    title = title.replace(' ', '_').replace('/', '_').replace('\\', '_')
    return f"{title}.xlsx"

# Write responses to an Excel file
def write_to_excel(data, title):
    script_path = os.path.dirname(os.path.abspath(__file__))
    data_path = os.path.abspath(os.path.join(script_path, '..', 'data'))

    file_name = generate_file_name(title)
    file_path = os.path.join(data_path, file_name)

    os.makedirs(data_path, exist_ok=True)

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Use the first entry to extract headers
    headers = data[0].keys()
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=i, value=header)

    for row_num, entry in enumerate(data, start=2):
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=row_num, column=col_num, value=entry.get(header))

    workbook.save(file_path)
    print(f"Data saved to {file_path}")
    return file_path

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    generate_prompt_path = os.path.join(script_dir, 'generate_prompt.py')
    
    # Call the generate_prompt.py script to generate the prompt
    if not os.path.exists(generate_prompt_path):
        print(f"Error: {generate_prompt_path} does not exist.")
        return

    subprocess.run(['python', generate_prompt_path])

    prompt_path = os.path.abspath(os.path.join(script_dir, '..', 'prompt.txt'))

    # Wait for the prompt file to be created by the GUI script
    while not os.path.exists(prompt_path):
        print("Waiting for prompt.txt to be generated...")
        time.sleep(1)

    with open(prompt_path, 'r') as file:
        prompt_content = file.read()

    # Extract the model and the prompt
    try:
        model_line = next(line for line in prompt_content.splitlines() if line.startswith("Model:"))
        model = model_line.split("Model: ")[1].strip()
        prompt = prompt_content.split("Task: ")[1].strip() 

        response = get_chatgpt_response(prompt, model)

    except Exception as e:
        print(f"Error parsing model or prompt: {e}")
        return

    try:
        json_part = extract_json_from_response(response)
        if json_part:
            response_list = json.loads(json_part)
            if isinstance(response_list, list) and len(response_list) > 0:
                title_entry = response_list[0]
                if "Title" in title_entry:
                    title = title_entry["Title"]
                    data = response_list[1:]
                    file_path = write_to_excel(data, title)

                    # Create a Tkinter root window (invisible) for the popup
                    root = tk.Tk()
                    root.withdraw()  
                    messagebox.showinfo("Success", f"Data successfully saved to {file_path}")
                    root.destroy()
                else:
                    print("Error: 'Title' field is missing in the response.")
            else:
                print("Error: Response is not a list or is empty.")
        else:
            print("Error: Could not extract JSON from the response.")
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON response: {e}")

if __name__ == "__main__":
    main()